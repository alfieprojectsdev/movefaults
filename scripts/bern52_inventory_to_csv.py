#!/usr/bin/env python3
"""
Export `docs/bern52/BERN52 Inventory.xlsx` to CSV, colours included.

WHY THIS EXISTS
---------------
In that workbook the **fill colour is the data**. Reading it the obvious
headless way -- ``openpyxl.load_workbook(..., read_only=True, data_only=True)``
-- discards styles and yields a mostly-empty sheet:

    sheet        cells with text    cells coloured with NO text
    Continuous            1,758                            727
    Campaign                442                          2,370

On the Campaign sheet roughly 84% of the information is fill colour. A value-only
read misses 3,097 cells and gives no sign that anything is wrong.

Two consequences the CSV fixes: the R740 can `grep` the inventory without
openpyxl, and `git diff` shows something readable instead of `Bin`.

THE TRAP THIS ENCODES
---------------------
**Each sheet has its own legend, and they disagree.** Blue means "Data
available" on Continuous and "No logsheets" on Campaign. The legends are
embedded in the sheets, so they are parsed from the swatch fills rather than
hardcoded -- if someone edits a legend, this follows.

There is also an undocumented near-duplicate: the Continuous body carries
``FF1E4E79`` (85 cells) alongside the legend's ``FF1F4E78`` (48 cells). One hex
digit apart in two channels, and the *undocumented* one is the larger group --
almost certainly the same category picked from a slightly different colour
picker. Unknown fills are matched to the nearest legend colour and flagged in
``colour_exact`` rather than silently dropped or silently merged.

USAGE
-----
    uv run python scripts/bern52_inventory_to_csv.py \\
        docs/bern52/"BERN52 Inventory.xlsx" \\
        docs/bern52/bern52_inventory.csv
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: uv add --optional dev openpyxl")

# Fills that are page furniture rather than categories: the column-background
# tints on the STATUS/TIME SERIES block, and the grey group-header bars.
FURNITURE = {"FFDDEBF7", "FFD9EAD3", "FFFCE4D6", "FF666666", "FFF3F3F3", "FFFFFFFF", "00000000"}

LEGEND_WORDS = (
    "data to be retrieved", "data available", "data unavailable", "old site",
    "occupied as campaign", "not operational", "downloaded data", "incomplete",
    "logsheets", "obsolete",
)


def _rgb(cell) -> str | None:
    f = cell.fill
    if f is None or f.fill_type != "solid" or f.fgColor is None:
        return None
    rgb = f.fgColor.rgb
    return rgb if isinstance(rgb, str) else None


def parse_legend(ws) -> dict[str, str]:
    """Read the legend embedded in the sheet: ``{RGB: meaning}``.

    Parsed rather than hardcoded so an edited legend is followed, and because
    the two sheets genuinely differ.
    """
    legend: dict[str, str] = {}
    for row in ws.iter_rows():
        texts = [str(c.value).strip() for c in row if c.value is not None and str(c.value).strip()]
        if not texts:
            continue
        joined = " | ".join(texts).lower()
        if not any(w in joined for w in LEGEND_WORDS):
            continue
        meaning = texts[-1]
        for c in row:
            rgb = _rgb(c)
            if rgb and rgb not in FURNITURE:
                legend[rgb] = meaning
    return legend


def _nearest(rgb: str, legend: dict[str, str]) -> tuple[str, bool]:
    """Exact legend match, else the closest legend colour, flagged inexact."""
    if rgb in legend:
        return legend[rgb], True
    if not legend:
        return "", False

    def chan(v: str) -> tuple[int, int, int]:
        return int(v[2:4], 16), int(v[4:6], 16), int(v[6:8], 16)

    r, g, b = chan(rgb)
    best = min(legend, key=lambda k: sum((a - c) ** 2 for a, c in zip(chan(k), (r, g, b), strict=True)))
    return legend[best], False


def export(xlsx: Path, out: Path) -> int:
    wb = openpyxl.load_workbook(xlsx)  # styles need read_only=False
    rows: list[dict] = []

    for ws in wb.worksheets:
        legend = parse_legend(ws)

        # Row 2 carries the year headers from column B; row 1 the trailing
        # metadata headers (STATUS, TIME SERIES, ...).
        years: dict[int, str] = {}
        for c in ws[2]:
            if c.column == 1 or c.value is None:
                continue
            m = re.match(r"^(\d{4})", str(c.value).strip())
            if m:
                years[c.column] = m.group(1)
        meta_cols = {
            c.column: str(c.value).strip()
            for c in ws[1]
            if c.value is not None and str(c.value).strip().isupper() and c.column > 2
        }

        group = ""
        for row in ws.iter_rows(min_row=3):
            site = row[0].value
            site = str(site).strip() if site is not None else ""

            # A grey bar with text but no site code is a group header.
            if not site:
                texts = [str(c.value).strip() for c in row if c.value is not None]
                if texts and any(_rgb(c) == "FF666666" for c in row):
                    group = texts[0]
                continue
            if len(site) != 4:
                continue

            # The SITE CODE CELL is itself coloured -- red for an old site,
            # yellow for "downloaded data this year". Dropping it would repeat
            # the exact failure this exporter exists to fix.
            site_rgb = _rgb(row[0])
            if site_rgb in FURNITURE:
                site_rgb = None
            site_meaning = _nearest(site_rgb, legend)[0] if site_rgb else ""

            meta = {
                name: (str(row[col - 1].value).strip() if col - 1 < len(row)
                       and row[col - 1].value is not None else "")
                for col, name in meta_cols.items()
            }

            for col, year in years.items():
                if col - 1 >= len(row):
                    continue
                cell = row[col - 1]
                text = str(cell.value).strip() if cell.value is not None else ""
                rgb = _rgb(cell)
                if rgb in FURNITURE:
                    rgb = None
                if not text and not rgb:
                    continue
                meaning, exact = _nearest(rgb, legend) if rgb else ("", True)
                rows.append(
                    {
                        "sheet": ws.title,
                        "group": group,
                        "site": site,
                        "year": year,
                        "doy_range": text,
                        "fill_rgb": rgb or "",
                        "colour_meaning": meaning,
                        "colour_exact": "" if not rgb else ("yes" if exact else "no"),
                        "site_fill_rgb": site_rgb or "",
                        "site_colour_meaning": site_meaning,
                        "status": meta.get("STATUS", ""),
                        "time_series": meta.get("TIME SERIES", ""),
                        "year_processed": meta.get("YEAR PROCESSED", ""),
                        "staff": meta.get("STAFF", ""),
                        "remarks": meta.get("REMARK/S", "") or meta.get("STATUS OF DATA", ""),
                    }
                )

    fields = [
        "sheet", "group", "site", "year", "doy_range", "fill_rgb",
        "colour_meaning", "colour_exact", "site_fill_rgb", "site_colour_meaning",
        "status", "time_series",
        "year_processed", "staff", "remarks",
    ]
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, lineterminator="\n")
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("csv_out", type=Path)
    a = ap.parse_args()
    n = export(a.xlsx, a.csv_out)
    print(f"wrote {n} rows -> {a.csv_out}")


if __name__ == "__main__":
    main()
